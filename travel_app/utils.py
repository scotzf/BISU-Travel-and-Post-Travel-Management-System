"""
Utility functions for travel app - HYBRID BUDGET EXTRACTION
Combines: Cell-based, Smart Regex, and Ollama AI
"""
import re
import tempfile
import os
from decimal import Decimal


def extract_budget_from_file(file):
    """
    HYBRID: Extract budget using multiple methods in priority order
    
    Args:
        file: Django UploadedFile object
        
    Returns:
        Decimal: Extracted budget amount or None if all methods fail
    """
    
    filename = file.name.lower()
    
    print(f"\n{'='*60}")
    print(f"🔍 BUDGET EXTRACTION: {filename}")
    print(f"{'='*60}")
    
    # METHOD 1: Cell-based (fastest, most accurate for standard layouts)
    print("\n[1/4] Trying cell-based extraction...")
    result = _try_cell_based_extraction(file, filename)
    if result:
        print(f"✅ SUCCESS via CELL-BASED: ₱{result:,.2f}")
        return result
    print("❌ Cell-based failed")
    
    # METHOD 2: Smart regex (fast, handles varied layouts)
    print("\n[2/4] Trying smart regex extraction...")
    result = _try_smart_regex_extraction(file, filename)
    if result:
        print(f"✅ SUCCESS via SMART REGEX: ₱{result:,.2f}")
        return result
    print("❌ Smart regex failed")
    
    # METHOD 3: Ollama AI (most flexible, works with any layout)
    print("\n[3/4] Trying Ollama AI extraction...")
    result = _try_ollama_extraction(file, filename)
    if result:
        print(f"✅ SUCCESS via OLLAMA AI: ₱{result:,.2f}")
        return result
    print("❌ Ollama AI failed")
    
    # METHOD 4: Fallback to date calculation
    print("\n[4/4] All extraction methods failed")
    print("⚠️ Will use date-based calculation as fallback")
    print(f"{'='*60}\n")
    return None


# ==================== METHOD 1: CELL-BASED ====================

def _try_cell_based_extraction(file, filename):
    """Extract from specific cells (works for standard layouts)"""
    
    if not filename.endswith(('.xlsx', '.xls')):
        return None
    
    try:
        import openpyxl
        
        # Reset file pointer
        file.seek(0)
        
        # CRITICAL: data_only=True to get calculated formula values
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
        
        # Strategy 1: Find "TOTAL" and get number in same row
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=100, values_only=False), 1):
            for col_idx, cell in enumerate(row, 1):
                if cell.value and isinstance(cell.value, str) and cell.value.strip().upper() == 'TOTAL':
                    # Look for number to the right (same row)
                    for offset in range(1, 15):
                        try:
                            next_cell = ws.cell(row=row_idx, column=col_idx + offset)
                            value = next_cell.value
                            
                            if isinstance(value, (int, float)) and value > 0:
                                print(f"   Found at Row {row_idx}, Col {col_idx + offset}")
                                return Decimal(str(value))
                        except:
                            continue
        
        # Strategy 2: Common cell locations
        common_cells = ['J40', 'K40', 'J14', 'K14', 'L40', 'M40']
        for cell_ref in common_cells:
            try:
                value = ws[cell_ref].value
                if isinstance(value, (int, float)) and value > 100:
                    print(f"   Found in cell {cell_ref}")
                    return Decimal(str(value))
            except:
                continue
        
    except Exception as e:
        print(f"   Error: {e}")
    
    return None


# ==================== METHOD 2: SMART REGEX ====================

def _try_smart_regex_extraction(file, filename):
    """Extract by reading ALL content and finding patterns"""
    
    # Reset file pointer
    file.seek(0)
    
    all_text = ""
    all_numbers = []
    
    # ===== EXCEL =====
    if filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            # Read EVERYTHING
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is not None:
                        all_text += f" {cell} "
                        if isinstance(cell, (int, float)):
                            all_numbers.append(float(cell))
            
        except Exception as e:
            print(f"   Error reading Excel: {e}")
            return None
    
    # ===== PDF =====
    elif filename.endswith('.pdf'):
        try:
            import pdfplumber
            with pdfplumber.open(file) as pdf:
                for page in pdf.pages[:5]:
                    all_text += page.extract_text() + " "
        except Exception as e:
            print(f"   Error reading PDF: {e}")
            return None
    
    # ===== WORD =====
    elif filename.endswith('.docx'):
        try:
            import docx
            doc = docx.Document(file)
            for paragraph in doc.paragraphs:
                all_text += paragraph.text + " "
        except Exception as e:
            print(f"   Error reading Word: {e}")
            return None
    
    # ===== IMAGE =====
    elif filename.endswith(('.jpg', '.jpeg', '.png', '.gif', '.bmp')):
        try:
            import easyocr
            
            # Save temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                file.seek(0)
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            
            reader = easyocr.Reader(['en'], gpu=False, verbose=False)
            result = reader.readtext(tmp_path)
            all_text = ' '.join([text[1] for text in result])
            
            os.unlink(tmp_path)
            
        except Exception as e:
            print(f"   Error with OCR: {e}")
            return None
    
    else:
        return None
    
    # Pattern matching with confidence scores
    patterns = [
        (r'total\s+estimated\s+(?:expenses?|budget|amount)[:\s]+₱?\s*([\d,]+\.?\d*)', 10),
        (r'total\s+(?:expenses?|budget|amount)[:\s]+₱?\s*([\d,]+\.?\d*)', 9),
        (r'grand\s+total[:\s]+₱?\s*([\d,]+\.?\d*)', 9),
        (r'estimated\s+total[:\s]+₱?\s*([\d,]+\.?\d*)', 8),
        (r'total[:\s]+₱?\s*([\d,]+\.?\d*)', 7),
        (r'budget[:\s]+₱?\s*([\d,]+\.?\d*)', 6),
    ]
    
    candidates = []
    
    for pattern, confidence in patterns:
        matches = re.finditer(pattern, all_text, re.IGNORECASE)
        for match in matches:
            try:
                amount = float(match.group(1).replace(',', ''))
                if 100 <= amount <= 1000000:
                    candidates.append((amount, confidence, pattern))
                    print(f"   Match: ₱{amount:,.2f} (confidence: {confidence})")
            except:
                pass
    
    # Return highest confidence match
    if candidates:
        candidates.sort(key=lambda x: x[1], reverse=True)
        return Decimal(str(candidates[0][0]))
    
    # Fallback: Find largest reasonable number
    reasonable = [n for n in all_numbers if 500 <= n <= 500000]
    if reasonable:
        largest = max(reasonable)
        print(f"   Fallback: Largest number ₱{largest:,.2f}")
        return Decimal(str(largest))
    
    return None


# ==================== METHOD 3: OLLAMA AI ====================

def _try_ollama_extraction(file, filename):
    """Use local Ollama AI to understand and extract budget"""
    
    try:
        import ollama
    except ImportError:
        print("   Ollama not installed (pip install ollama)")
        return None
    
    # Reset file pointer
    file.seek(0)
    
    # Step 1: Get file content
    content = ""
    
    if filename.endswith(('.xlsx', '.xls')):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
            
            rows = []
            for row in ws.iter_rows(values_only=True, max_row=100):
                row_text = ' | '.join([str(cell) if cell else '' for cell in row])
                if row_text.strip():
                    rows.append(row_text)
            
            content = '\n'.join(rows)
        except Exception as e:
            print(f"   Error reading Excel: {e}")
            return None
    
    elif filename.endswith(('.jpg', '.jpeg', '.png')):
        try:
            import easyocr
            
            with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp:
                file.seek(0)
                for chunk in file.chunks():
                    tmp.write(chunk)
                tmp_path = tmp.name
            
            reader = easyocr.Reader(['en'], gpu=False, verbose=False)
            result = reader.readtext(tmp_path)
            content = '\n'.join([text[1] for text in result])
            
            os.unlink(tmp_path)
        except Exception as e:
            print(f"   Error with OCR: {e}")
            return None
    
    elif filename.endswith('.pdf'):
        try:
            import pdfplumber
            with pdfplumber.open(file) as pdf:
                content = '\n'.join([page.extract_text() for page in pdf.pages[:5]])
        except Exception as e:
            print(f"   Error reading PDF: {e}")
            return None
    
    if not content:
        return None
    
    # Step 2: Ask Ollama
    try:
        print("   Querying Ollama AI...")
        
        response = ollama.chat(
            model='llama3.2:3b',
            messages=[{
                'role': 'user',
                'content': f'''You are analyzing a Philippine travel itinerary document.
Find the TOTAL estimated budget or total estimated expenses amount.

Document content:
{content[:3000]}  

Return ONLY the numerical amount (no peso sign, no commas).
If you cannot find a clear total, return: NOT_FOUND

Examples:
- "Total: ₱1,234.56" → return: 1234.56
- "Total Estimated: 5000" → return: 5000
- If unclear → return: NOT_FOUND'''
            }],
            options={
                'temperature': 0,  # More deterministic
                'num_predict': 50  # Short response
            }
        )
        
        answer = response['message']['content'].strip()
        print(f"   AI Response: {answer}")
        
        if answer != 'NOT_FOUND':
            # Clean and parse
            answer = answer.replace(',', '').replace('₱', '').strip()
            try:
                amount = float(answer)
                if 100 <= amount <= 1000000:
                    return Decimal(str(amount))
            except:
                pass
        
    except Exception as e:
        print(f"   Ollama error: {e}")
    
    return None


# ==================== FALLBACK CALCULATOR ====================

def calculate_auto_budget(start_date, end_date, region_rate):
    """
    Fallback: Calculate budget based on dates
    
    Args:
        start_date: Date object
        end_date: Date object
        region_rate: RegionRate object
        
    Returns:
        Decimal: Calculated budget
    """
    duration_days = (end_date - start_date).days + 1
    duration_nights = max(0, (end_date - start_date).days)
    
    meals_cost = duration_days * 3 * region_rate.meal_rate
    lodging_cost = duration_nights * region_rate.lodging_rate
    incidental_cost = duration_days * region_rate.incidental_rate
    
    total = meals_cost + lodging_cost + incidental_cost
    
    print(f"\n📊 Auto-calculated budget:")
    print(f"   Meals: {duration_days} days × 3 × ₱{region_rate.meal_rate} = ₱{meals_cost}")
    print(f"   Lodging: {duration_nights} nights × ₱{region_rate.lodging_rate} = ₱{lodging_cost}")
    print(f"   Incidentals: {duration_days} days × ₱{region_rate.incidental_rate} = ₱{incidental_cost}")
    print(f"   TOTAL: ₱{total}")
    
    return Decimal(str(total))